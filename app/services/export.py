import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Tuple
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.db import Transaction, Goal


async def generate_csv_exports(session: AsyncSession) -> Tuple[bytes, bytes]:
    """
    Generates CSV export bytes for transactions and goals.
    Returns (transactions_csv_bytes, goals_csv_bytes).
    """
    # 1. Export Transactions
    tx_stmt = select(Transaction).order_by(Transaction.date.desc(), Transaction.id.desc())
    tx_res = await session.execute(tx_stmt)
    transactions = tx_res.scalars().all()

    tx_output = io.StringIO()
    tx_writer = csv.writer(tx_output)
    tx_writer.writerow(["id", "author_telegram_id", "type", "amount", "currency", "category", "note", "date", "source", "created_at"])
    for tx in transactions:
        tx_writer.writerow([
            tx.id,
            tx.author_telegram_id,
            tx.type,
            str(tx.amount),
            tx.currency,
            tx.category or "",
            tx.note or "",
            tx.date.isoformat(),
            tx.source,
            tx.created_at.isoformat()
        ])

    # 2. Export Goals
    goal_stmt = select(Goal).order_by(Goal.id.asc())
    goal_res = await session.execute(goal_stmt)
    goals = goal_res.scalars().all()

    goal_output = io.StringIO()
    goal_writer = csv.writer(goal_output)
    goal_writer.writerow(["id", "title", "target_amount", "current_amount", "currency", "deadline", "apy", "monthly_contribution_plan", "status", "created_at"])
    for g in goals:
        goal_writer.writerow([
            g.id,
            g.title,
            str(g.target_amount),
            str(g.current_amount),
            g.currency,
            g.deadline.isoformat() if g.deadline else "",
            g.apy if g.apy is not None else "",
            str(g.monthly_contribution_plan) if g.monthly_contribution_plan else "",
            g.status,
            g.created_at.isoformat()
        ])

    tx_bytes = tx_output.getvalue().encode("utf-8-sig")
    goal_bytes = goal_output.getvalue().encode("utf-8-sig")

    return tx_bytes, goal_bytes

async def generate_excel_exports(session: AsyncSession) -> bytes:
    """
    Generates Excel export bytes for transactions and goals.
    """
    wb = openpyxl.Workbook()
    
    # 1. Export Transactions
    ws_tx = wb.active
    ws_tx.title = "Транзакции"
    headers_tx = ["ID", "Дата", "Тип", "Сумма", "Валюта", "Категория", "Заметка", "Источник"]
    ws_tx.append(headers_tx)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num in range(1, len(headers_tx) + 1):
        cell = ws_tx.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    tx_stmt = select(Transaction).order_by(Transaction.date.desc(), Transaction.id.desc())
    tx_res = await session.execute(tx_stmt)
    transactions = tx_res.scalars().all()
    
    for tx in transactions:
        ws_tx.append([
            tx.id,
            tx.date.strftime("%d.%m.%Y"),
            tx.type,
            float(tx.amount),
            tx.currency,
            tx.category or "",
            tx.note or "",
            tx.source
        ])
        
    ws_tx.column_dimensions["B"].width = 12
    ws_tx.column_dimensions["D"].width = 15
    ws_tx.column_dimensions["F"].width = 20
    ws_tx.column_dimensions["G"].width = 30
    
    # 2. Export Goals
    ws_goal = wb.create_sheet("Цели")
    headers_goal = ["ID", "Название", "Цель", "Накоплено", "Валюта", "Дедлайн", "Статус"]
    ws_goal.append(headers_goal)
    
    for col_num in range(1, len(headers_goal) + 1):
        cell = ws_goal.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        
    goal_stmt = select(Goal).order_by(Goal.id.asc())
    goal_res = await session.execute(goal_stmt)
    goals = goal_res.scalars().all()
    
    for g in goals:
        ws_goal.append([
            g.id,
            g.title,
            float(g.target_amount),
            float(g.current_amount),
            g.currency,
            g.deadline.strftime("%d.%m.%Y") if g.deadline else "",
            g.status
        ])
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
